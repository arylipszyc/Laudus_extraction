"""Tests del reporte de gastos para contadores — Story 10.2 (guard cuentas sin categorizar).

Reproduce el defecto verificado: una cuenta creada nueva en Laudus (ausente del plan de
cuentas → Categoria vacía) hoy desaparece del reporte en silencio. El guard debe hacerla
visible y, si su prefijo la ubica en un centro de costo de gasto conocido, contarla en
TOTAL EGRESOS.
"""
import io
from datetime import date

from openpyxl import load_workbook

from backend.app.api.v1.reportes.report_builder import build_report


def _ws(rows, start=date(2025, 3, 1), end=date(2025, 3, 31)):
    out = build_report(start, end, lambda _sheet: rows)
    return load_workbook(io.BytesIO(out)).active


def _labels(ws):
    return {(c.value or ""): c.row for c in ws["A"] if c.value}


def _appears(ws, amount):
    target = round(amount)
    return any(
        isinstance(c.value, (int, float)) and round(c.value) == target
        for row in ws.iter_rows()
        for c in row
    )


def test_cuenta_nueva_con_prefijo_conocido_aparece_y_suma_a_egresos():
    """413xxx = Casa Sur. Una cuenta nueva sin categoría debe aparecer y entrar al total."""
    rows = [
        {"date": "2025-03-10", "accountnumber": "413077", "accountName": "GASTO NUEVO CASA SUR",
         "debit": "9999999", "credit": "0", "Categoria1": "", "Categoria2": "", "Categoria3": ""},
    ]
    ws = _ws(rows)

    # Antes del guard: NO aparecía en ninguna celda. Ahora debe aparecer.
    assert _appears(ws, 9999999), "la cuenta nueva sin categoría no aparece en el reporte"

    labels = _labels(ws)
    # Hay una fila de control en el resumen, incluida en TOTAL EGRESOS.
    sincat_label = next((l for l in labels if "sin categorizar" in str(l).lower()), None)
    assert sincat_label is not None, "falta la fila 'sin categorizar' en el resumen"
    sincat_row = labels[sincat_label]

    total_label = next(l for l in labels if str(l).strip().upper() == "TOTAL EGRESOS")
    total_formula = ws.cell(labels[total_label], 2).value
    assert f"B{sincat_row}" in str(total_formula), \
        f"TOTAL EGRESOS ({total_formula}) no incluye la fila sin categorizar (B{sincat_row})"

    # Y aparece itemizada con su nombre y código en la sección de detalle.
    assert any("GASTO NUEVO CASA SUR" in str(l) and "413077" in str(l) for l in labels), \
        "la cuenta nueva no está itemizada en la sección CUENTAS SIN CATEGORIZAR"


def test_prefijo_desconocido_aparece_pero_no_suma():
    """Prefijo de familia gasto (8xx) pero no reconocido: visible, NO sumado a ningún total."""
    rows = [
        {"date": "2025-03-10", "accountnumber": "888001", "accountName": "GASTO RARO",
         "debit": "500000", "credit": "0", "Categoria1": "", "Categoria2": "", "Categoria3": ""},
    ]
    ws = _ws(rows)
    assert _appears(ws, 500000), "la cuenta de prefijo desconocido no aparece"
    labels = _labels(ws)
    assert any("desconocido" in str(l).lower() for l in labels), \
        "no se marca el caso de prefijo desconocido"
    # No debe haber fila 'sin categorizar' en el resumen (no se suma a EGRESOS).
    total_label = next(l for l in labels if str(l).strip().upper() == "TOTAL EGRESOS")
    total_formula = str(ws.cell(labels[total_label], 2).value)
    # El total solo debe referenciar eag_sub + hijas_sub (2 términos), sin la cuenta rara.
    assert total_formula.count("+") <= 1, \
        f"TOTAL EGRESOS no debería sumar el prefijo desconocido: {total_formula}"


def test_codigo_no_parseable_sin_categoria_sigue_visible():
    """Code-review patch 1: un código no numérico sin categoría no debe caerse del reporte."""
    rows = [
        {"date": "2025-03-10", "accountnumber": "VARIOS", "accountName": "GASTO RARO SIN CODIGO",
         "debit": "777777", "credit": "0", "Categoria1": "", "Categoria2": "", "Categoria3": ""},
    ]
    ws = _ws(rows)
    assert _appears(ws, 777777), "una cuenta sin categoría y con código no parseable se perdió"
    labels = _labels(ws)
    assert any("desconocido" in str(l).lower() for l in labels)


def test_cuentas_categorizadas_no_disparan_el_guard():
    """No-regresión: con cuentas bien categorizadas no aparece sección de sin-categorizar."""
    rows = [
        {"date": "2025-03-10", "accountnumber": "413005", "accountName": "LUZ",
         "debit": "1000", "credit": "0", "Categoria1": "GASTOS - EGRESOS",
         "Categoria2": "Casa Sur", "Categoria3": "x"},
    ]
    ws = _ws(rows)
    labels = _labels(ws)
    assert not any("sin categorizar" in str(l).lower() for l in labels), \
        "el guard no debe disparar con data categorizada"
    assert not any("desconocido" in str(l).lower() for l in labels)
