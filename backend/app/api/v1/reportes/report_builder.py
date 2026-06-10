"""Construye el reporte de gastos desde Laudus, con la estructura del Excel original.

Secciones: INGRESOS · EGRESOS (resumen) · DETALLE DE LOS GASTOS · TARJETAS DE CRÉDITO.
El orden/estructura de INGRESOS y DETALLE se toma de `template_order.json`. Cada línea
se resuelve por código de cuenta contra `ledger_final`. Subtotales y totales se escriben
como FÓRMULAS Excel (=SUM...) para que se recalculen cuando el contador complete las
celdas vacías. Subtotales en gris claro, totales en gris oscuro.
"""
from __future__ import annotations

import io
import json
import os
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LIGHT = PatternFill("solid", fgColor="D9D9D9")   # subtotales
DARK = PatternFill("solid", fgColor="A6A6A6")    # totales
BOLD = Font(bold=True)
TITLE = Font(bold=True, size=12)
MONEY = "#,##0;-#,##0"                            # negativos sin rojo
BOTTOM = Border(bottom=Side(style="thin", color="808080"))

TEMPLATE = os.path.join(os.path.dirname(__file__), "template_order.json")

# El template es estático: se carga y particiona una sola vez al importar el módulo,
# no en cada request a GET /reportes/gastos.
with open(TEMPLATE, encoding="utf-8") as _f:
    _TEMPLATE = json.load(_f)
ING_TPL = [e for e in _TEMPLATE if e["section"] == "INGRESOS"]
EGR_TPL = [e for e in _TEMPLATE if e["section"] == "EGRESOS"]

ALIAS = {
    "310025": ["113002", "113003"],
    "6900001": ["690001"], "69000003": ["690003"], "6501073": ["690073"], "6900099": ["690099"],
    "79000003": ["790003"], "7900099": ["790099"],
    "9900001": ["990001"], "9900003": ["990003"], "9501073": ["990073"], "9900099": ["990099"],
}
LEVEL = {
    "TOTAL CUENTAS BASICAS": "group", "TOTAL SUELDOS BRUTOS": "group",
    "TOTAL VEHICULOS": "group", "TOTAL VARIOS": "group",
    "TOTAL DEPTO SANTIAGO": "cc", "TOTAL CASA SUR": "cc",
    "TOTAL DEPARTAMENTO MIAMI": "cc", "TOTAL GASTOS PERSONALES EAG": "cc",
    "TOTAL INGRESOS EAG": "cc",
    "TOTAL GASTOS JOCELYN": "cc_daughter", "TOTAL GASTOS JEANNETTE": "cc_daughter",
    "TOTAL GASTOS JOHANNA": "cc_daughter", "TOTAL GASTOS JAEL": "cc_daughter",
    "TOTAL GASTOS HIJAS": "hijas", "SUBTOTAL INGRESOS HIJAS": "hijas",
    "TOTAL INGRESOS": "grand",
}
GRAND = {"TOTAL INGRESOS", "TOTAL EGRESOS"}        # gris oscuro
BLANK_AFTER = {"TOTAL INGRESOS EAG", "TOTAL GASTOS PERSONALES EAG"}  # fila de separación EAG/Hijas
DAUGHTERS = [
    ("Egresos Jocelyn", "EGRESOS JOCELYN AVAYU DEUTSCH"),
    ("Egresos Jeannette", "EGRESOS JEANNETTE AVAYU DEUTSCH"),
    ("Egresos Johanna", "EGRESOS JOHANNA AVAYU DEUTSCH"),
    ("Egresos Jael", "EGRESOS JAEL AVAYU DEUTSCH"),
]

# Story 10.2 — guard "cuentas sin categorizar".
# Una cuenta creada nueva en Laudus (ausente del plan de cuentas) llega con Categoria vacía
# y hoy desaparece del reporte: el resumen agrupa por Categoria2 (cae en el bucket "" que nunca
# se lee) y el detalle exige el código en template_order.json. Detectamos esas cuentas por su
# PREFIJO de número de cuenta. Semilla con los 4 CC de gasto EAG (prefijo→CC verificado 1:1 contra
# Laudus); los prefijos de hijas se aprenden de la data ya categorizada (ver build_report).
EAG_EXPENSE_PREFIX = {
    "411": "Departamento Santiago", "413": "Casa Sur",
    "415": "Departamento Miami", "430": "Gastos Personales",
}
KNOWN_EAG_CAT2 = {"DEPARTAMENTO SANTIAGO", "Casa Sur", "DEPARTAMENTO MIAMI", "GASTOS PERSONALES"}
EXPENSE_FIRST_DIGITS = set("46789")  # familias de gasto del plan (4xx EAG, 6/7/8/9xx hijas)


def _cc_prefix(code):
    """Prefijo de 3 dígitos del número de cuenta (4101→411), o "" si no aplica."""
    s = str(code).strip().replace(" ", "").replace("-", "")
    if s.startswith("4101"):
        s = "411" + s[4:]
    return s[:3] if s[:3].isdigit() else ""


def _num(v):
    try:
        return float(str(v).replace(",", "") or 0)
    except (ValueError, TypeError):
        return 0.0


def _months_in_range(start, end):
    out, y, m = [], start.year, start.month
    while (y, m) <= (end.year, end.month):
        out.append((y, m))
        m, y = (1, y + 1) if m == 12 else (m + 1, y)
    return out


def norm_codes(a):
    s = str(a).strip().upper()
    if not s or s.startswith("TC") or s.startswith("T/C"):
        return []
    s = s.replace(" ", "").replace("-", "")
    parts = s.split("/")
    if not parts[0].isdigit():
        return []
    base = parts[0]
    if base.startswith("4101"):
        base = "411" + base[4:]
    codes = [base] + [base[: len(base) - len(p)] + p for p in parts[1:] if p.isdigit()]
    out = []
    for c in codes:
        out.extend(ALIAS.get(c, [c]))
    return out


def _laudus_by_code(rows, months):
    mindex = {ym: i for i, ym in enumerate(months)}
    is_income = defaultdict(bool)
    for r in rows:
        if "INGRESOS" in str(r.get("Categoria1", "")).upper():
            is_income[str(r.get("accountnumber", ""))] = True
    idx = defaultdict(lambda: [0.0] * len(months))
    for r in rows:
        ds = str(r.get("date", ""))[:10]
        if len(ds) < 7:
            continue
        mi = mindex.get((int(ds[:4]), int(ds[5:7])))
        if mi is None:
            continue
        acc = str(r.get("accountnumber", ""))
        de, cr = _num(r.get("debit")), _num(r.get("credit"))
        idx[acc][mi] += (cr - de) if is_income[acc] else (de - cr)
    return idx


def build_report(start: date, end: date, get_records) -> bytes:
    rows = get_records("ledger_final")
    months = _months_in_range(start, end)
    nmon = len(months)
    laudus = _laudus_by_code(rows, months)
    have = {c for c in laudus if any(laudus[c])}
    used = set()
    tc_codes = {str(r.get("accountnumber", "")) for r in rows
                if str(r.get("Categoria2", "")) == "GASTOS PERSONALES"
                and str(r.get("Categoria3", "")) == "Tarjetas Credito"}

    mindex = {ym: i for i, ym in enumerate(months)}
    cat2 = defaultdict(lambda: [0.0] * nmon)
    cat1 = defaultdict(lambda: [0.0] * nmon)
    for r in rows:
        ds = str(r.get("date", ""))[:10]
        if len(ds) < 7:
            continue
        mi = mindex.get((int(ds[:4]), int(ds[5:7])))
        if mi is None:
            continue
        amt = _num(r.get("debit")) - _num(r.get("credit"))
        cat2[str(r.get("Categoria2", ""))][mi] += amt
        cat1[str(r.get("Categoria1", ""))][mi] += amt

    # ----- Story 10.2: detección de cuentas sin categorizar -----
    name_of, acc_cat = {}, {}
    for row in rows:
        acc = str(row.get("accountnumber", ""))
        if acc and acc not in acc_cat:
            acc_cat[acc] = (str(row.get("Categoria1", "")), str(row.get("Categoria2", "")))
            name_of[acc] = str(row.get("accountName", "")) or acc
    daughter_cat1 = {k for _, k in DAUGHTERS}
    prefix_label = dict(EAG_EXPENSE_PREFIX)        # semilla EAG + prefijos de hijas aprendidos de la data
    for acc, (c1, c2) in acc_cat.items():
        if c2 in KNOWN_EAG_CAT2:
            prefix_label.setdefault(_cc_prefix(acc), c2)
        elif c1 in daughter_cat1:
            prefix_label.setdefault(_cc_prefix(acc), c1)
    # El defecto es una cuenta SIN categoría (nueva en Laudus, ausente del plan). Una cuenta con
    # categoría —aunque no sea de gasto, p.ej. un activo "DISPONIBLE…"— está bien y NO se marca.
    uncat, uncat_unknown = [], []   # (code, name, cc_label, values) / (code, name, values)
    for acc in sorted(have):
        c1, c2 = acc_cat.get(acc, ("", ""))
        if c1.strip() or c2.strip():
            continue
        p = _cc_prefix(acc)
        if p in prefix_label:                       # prefijo ubica un CC de gasto → cuenta al total
            uncat.append((acc, name_of.get(acc, acc), prefix_label[p], laudus[acc]))
        elif p[:1] in EXPENSE_FIRST_DIGITS or not p:  # parece gasto, o código no parseable → visible, no sumado
            uncat_unknown.append((acc, name_of.get(acc, acc), laudus[acc]))
    uncat_total = [sum(v[3][m] for v in uncat) for m in range(nmon)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws.sheet_view.showGridLines = False
    total_col = 2 + nmon
    LC = [get_column_letter(2 + j) for j in range(nmon)]   # letras de columnas de mes
    r = 1

    def row_total_formula(row):
        return f"=SUM({LC[0]}{row}:{LC[-1]}{row})" if nmon else None

    def write(label, *, values=None, month_formulas=None, total_formula=None,
              bold=False, indent=False, fill=None):
        nonlocal r
        cur = r
        cell = ws.cell(cur, 1, ("   " if indent else "") + label)
        if bold:
            cell.font = BOLD
        for j in range(nmon):
            c = ws.cell(cur, 2 + j)
            if month_formulas is not None:
                c.value = month_formulas[j]
            elif values is not None:
                c.value = round(values[j])
            c.number_format = MONEY
            if bold:
                c.font = BOLD
        tc = ws.cell(cur, total_col)
        tc.value = total_formula if total_formula is not None else row_total_formula(cur)
        tc.number_format = MONEY
        if bold:
            tc.font = BOLD
        if fill:
            for c in range(1, total_col + 1):
                ws.cell(cur, c).fill = fill
        r += 1
        return cur

    def section_title(t):
        nonlocal r
        r += 1
        ws.cell(r, 1, t).font = TITLE
        r += 1

    def month_header():
        nonlocal r
        for c, txt in [(1, "Detalle")] + [(2 + j, f"{m:02d}/{y}") for j, (y, m) in enumerate(months)] + [(total_col, "TOTAL")]:
            cell = ws.cell(r, c, txt)
            cell.font, cell.border = BOLD, BOTTOM
            if c > 1:
                cell.alignment = Alignment(horizontal="center")
        r += 1

    def sub_fill(label):
        return DARK if label.strip().upper() in GRAND else LIGHT

    def cells_formula(child_rows):
        return ["=" + "+".join(f"{LC[j]}{rr}" for rr in child_rows) for j in range(nmon)]

    def range_formula(first, last):
        return [f"=SUM({LC[j]}{first}:{LC[j]}{last})" for j in range(nmon)]

    def render_template(entries):
        nonlocal r
        det_first = det_last = None
        grp_rows, daughter_rows, sec_rows = [], [], []
        for e in entries:
            if e["kind"] == "header":
                ws.cell(r, 1, e["label"]).font = BOLD
                r += 1
            elif e["kind"] == "detail":
                codes = [c for c in norm_codes(e["code"]) if c in have and c not in used]
                if codes:
                    vals = [sum(laudus[c][m] for c in codes) for m in range(nmon)]
                    row = write(e["label"], values=vals, indent=True)
                    used.update(codes)
                else:
                    row = write(e["label"], indent=True)  # vacía, a completar
                det_first = det_first or row
                det_last = row
            elif e["kind"] == "subtotal":
                lbl = e["label"].strip().upper()
                if lbl == "TOTAL GASTOS PERSONALES EAG":
                    fresh = [c for c in tc_codes if c in have and c not in used]
                    if fresh:
                        vals = [sum(laudus[c][m] for c in fresh) for m in range(nmon)]
                        row = write("Tarjetas de Crédito (total Laudus — a desglosar)",
                                    values=vals, indent=True)
                        used.update(fresh)
                        det_first = det_first or row
                        det_last = row
                lvl = LEVEL.get(lbl, "group")
                if lvl in ("group", "cc_daughter") or (lvl == "cc" and not grp_rows):
                    mf = range_formula(det_first, det_last) if det_first else None
                elif lvl == "cc":
                    mf = cells_formula(grp_rows)
                elif lvl == "hijas":
                    mf = cells_formula(daughter_rows) if daughter_rows else (
                        range_formula(det_first, det_last) if det_first else None)
                else:  # grand
                    mf = cells_formula(sec_rows)
                row = write(e["label"], month_formulas=mf, bold=True, fill=sub_fill(lbl))
                if lvl == "group":
                    grp_rows.append(row)
                elif lvl == "cc":
                    sec_rows.append(row); grp_rows = []
                elif lvl == "cc_daughter":
                    daughter_rows.append(row)
                elif lvl == "hijas":
                    sec_rows.append(row); daughter_rows = []
                det_first = det_last = None
                if lbl in BLANK_AFTER:
                    r += 1  # fila de separación EAG / Hijas

    # ----- Encabezado -----
    ws.cell(r, 1, "REPORTE DE GASTOS — EAG").font = Font(bold=True, size=14); r += 1
    ws.cell(r, 1, f"Período: {start.isoformat()} a {end.isoformat()}  ·  Fuente: Laudus (ledger_final)"); r += 1
    ws.cell(r, 1, "Celdas vacías = a completar por el contador. Subtotales/totales son fórmulas.").font = Font(italic=True, size=9)
    r += 1

    # ----- INGRESOS -----
    section_title("INGRESOS")
    month_header()
    render_template(ING_TPL)

    # ----- EGRESOS (resumen) -----
    section_title("EGRESOS (resumen)")
    month_header()
    ws.cell(r, 1, "EGRESOS EAG").font = BOLD; r += 1
    cc_rows = []
    for label, key in [("Departamento Santiago", "DEPARTAMENTO SANTIAGO"), ("Casa Sur", "Casa Sur"),
                       ("Departamento Miami", "DEPARTAMENTO MIAMI"), ("Gastos Personales", "GASTOS PERSONALES")]:
        cc_rows.append(write(label, values=cat2[key], indent=True))
    eag_sub = write("Subtotal Egresos EAG", month_formulas=cells_formula(cc_rows), bold=True, fill=LIGHT)
    r += 1  # separación EAG / Hijas
    ws.cell(r, 1, "EGRESOS HIJAS").font = BOLD; r += 1
    h_rows = [write(label, values=cat1[key], indent=True) for label, key in DAUGHTERS]
    hijas_sub = write("Subtotal Egresos Hijas", month_formulas=cells_formula(h_rows), bold=True, fill=LIGHT)
    total_terms = [eag_sub, hijas_sub]
    if uncat:  # Story 10.2: las cuentas sin categorizar (CC conocido) entran al total (consistente con el detalle)
        total_terms.append(
            write("⚠️ Egresos sin categorizar (ver detalle al final)", values=uncat_total, indent=True))
    write("TOTAL EGRESOS", month_formulas=cells_formula(total_terms), bold=True, fill=DARK)

    # ----- DETALLE DE LOS GASTOS -----
    section_title("DETALLE DE LOS GASTOS")
    month_header()
    render_template(EGR_TPL)

    # ----- TARJETAS DE CRÉDITO (a completar por el contador) -----
    section_title("TARJETAS DE CRÉDITO — detalle a completar por el contador (desde cartola)")
    month_header()
    tc_lump = [sum(laudus[c][m] for c in tc_codes if c in laudus) for m in range(nmon)]
    write("Total según Laudus (objetivo a desglosar)", values=tc_lump, bold=True, fill=LIGHT)

    def _resolvable(code):
        return any(c in have for c in norm_codes(code))

    cats, ingp = [], False
    for e in EGR_TPL:
        if e["kind"] == "header" and e["label"].strip().upper() == "GASTOS PERSONALES":
            ingp = True; continue
        if e["kind"] == "subtotal" and "GASTOS PERSONALES" in e["label"].upper():
            ingp = False
        if ingp and e["kind"] == "detail" and not _resolvable(e["code"]):
            cats.append(e["label"])
    cat_first = r
    for c in cats:
        write(c, indent=True)  # fila vacía a completar
    if r - 1 >= cat_first:
        write("Subtotal desglose (debe igualar el total Laudus de arriba)",
              month_formulas=range_formula(cat_first, r - 1), bold=True, fill=LIGHT)

    # ----- Story 10.2: CUENTAS SIN CATEGORIZAR -----
    if uncat or uncat_unknown:
        section_title("⚠️ CUENTAS SIN CATEGORIZAR — en Laudus pero faltan en el plan de cuentas")
        month_header()
        for code, name, cc_label, vals in uncat:
            write(f"{name} · {cc_label} · {code}", values=vals, indent=True)
        if uncat:
            write("Total sin categorizar (incluido en TOTAL EGRESOS)",
                  values=uncat_total, bold=True, fill=LIGHT)
        for code, name, vals in uncat_unknown:
            write(f"{name} · (centro de costo desconocido — revisar) · {code}",
                  values=vals, indent=True)
        if uncat_unknown:
            unk_total = [sum(v[2][m] for v in uncat_unknown) for m in range(nmon)]
            write("Total prefijo desconocido (NO incluido en ningún total — categorizar)",
                  values=unk_total, bold=True, fill=LIGHT)

    ws.column_dimensions["A"].width = 46
    for j in range(nmon + 1):
        ws.column_dimensions[get_column_letter(2 + j)].width = 13

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
